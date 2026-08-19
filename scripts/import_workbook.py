"""Convert the supplied BDSP checklist workbook into static TypeScript data.

The generated file is intentionally committed as app data so the deployed SPA
does not need Python, Excel, or a backend. Re-run this script when the source
workbook is updated.
"""
from __future__ import annotations

import json
import re
import sys
import urllib.request
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from lxml import html as lxml_html

import openpyxl

SOURCE = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(r"C:\Users\Usuario\Downloads\Cópia de BDSP Pokedex Worklist Sharable.xlsx")
ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "src" / "data" / "gameData.ts"
API_CACHE = ROOT / ".cache" / "pokeapi"
POKEMONDB_CACHE = ROOT / ".cache" / "pokemondb"


def clean(value) -> str:
    return str(value or "").replace("\r", "").strip()


def slug(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")


def location_parts(raw: str):
    for line in clean(raw).split("\n"):
        line = line.strip()
        if not line:
            continue
        bits = [line]
        if " / " in line and "PokeRadar" in line:
            bits = [part.strip() + " - PokeRadar" for part in line.split(" / ")]
        elif line.lower() in {"great marsh, trophy garden", "great marsh, trophy garden - post game"}:
            bits = [part.strip() for part in line.split(",")]
        for bit in bits:
            yield bit


def normalize_location(raw: str):
    original = raw.strip().replace('[[', '').replace(']]', '')
    original = re.sub(r"\bRote\s+(\d+)", r"Route \1", original, flags=re.I)
    # Correct three route labels present in the workbook with missing/swapped
    # digits. Keep this in the import boundary so the generated app data and
    # future re-imports remain consistent with Sinnoh's Route 201-230 map.
    original = {
        "Route 121": "Route 212",
        "Route 22": "Route 211",
        "Route 30": "Route 230",
    }.get(original, original)
    original = re.sub(r"\bGrand Undeground\b", "Grand Underground", original, flags=re.I)
    original = re.sub(r"\bSunshore City\b", "Sunyshore City", original, flags=re.I)
    original = re.sub(r"\bMount Coronet\b", "Mt. Coronet", original, flags=re.I)
    lower = original.lower()
    condition = None
    method = None
    times = []
    if 'morning only' in lower:
        times = ['morning']
    elif 'night only' in lower:
        times = ['night']
    elif 'day only' in lower:
        times = ['day']
    if "pokeradar" in lower or "poke radar" in lower:
        method = "PokéRadar"
    elif "super rod" in lower:
        method = "Super Rod"
    elif "old rod" in lower:
        method = "Old Rod"
    elif "good rod" in lower:
        method = "Good Rod"
    elif re.search(r"\bsurf\b", lower):
        method = "Surf"
    elif "honey" in lower:
        method = "Honey Tree"
    elif "swarm" in lower:
        method = "Swarm"
    if "post game" in lower:
        condition = "Pós-jogo"
    elif "late game" in lower:
        condition = "Late game"

    base = re.sub(r"\s*-\s*(PokeRadar|Poke Radar|Super Rod|Old Rod|Good Rod|Surf|Swarm(?: Pokemon)?)(\s*-\s*(Post Game|Late Game))?", "", original, flags=re.I).strip()
    base = re.sub(r"\s*\((Morning|Night|Day) Only\)", "", base, flags=re.I).strip()
    base = re.sub(r"\s*-\s*(Post Game|Late Game)", "", base, flags=re.I).strip()
    if base.lower().startswith("grand underground"):
        base = "Grand Underground"
    if base.lower().startswith("evolve "):
        condition = base
        base = "Evolution"
    if base.lower().startswith("starter pokemon"):
        condition = "Starter Pokémon"
        base = "Starter Pokémon"
    if re.match(r"^most water\b", base, re.I):
        base = "Open water"
    if base.lower() in {"old rod", "good rod", "super rod"}:
        base = "Open water"
    if base.lower() == "etc.":
        base = "Other routes"
    category = "route" if re.match(r"Route \d+", base, re.I) else ("special" if base in {"Evolution", "Starter Pokémon", "Grand Underground", "Ramanas Park", "Honey Trees", "Trophy Garden", "Great Marsh", "Open water"} else "area")
    return base, category, method, condition, times


def acquisition_group(label: str):
    value = label.lower()
    if "breed" in value: return "breed"
    if any(token in value for token in ("mystery gift", "event", "glitch")): return "event"
    if "trade" in value: return "trade"
    if any(token in value for token in ("egg", "starter", "gift")): return "egg"
    if "evolve" in value or "shedinja" in value: return "evolution"
    return "special"


def is_acquisition(label: str):
    value = label.lower()
    return any(token in value for token in ("breed", "mystery gift", "event", "glitch", "trade", "egg", "starter", "gift", "evolve", "shedinja", "revive", "see how to get", "future", "currently uknown", "unknown if"))


def api_encounters(pokemon_id: int):
    API_CACHE.mkdir(parents=True, exist_ok=True)
    cache = API_CACHE / f"{pokemon_id}.json"
    if cache.exists():
        return json.loads(cache.read_text(encoding="utf-8"))
    request = urllib.request.Request(f"https://pokeapi.co/api/v2/pokemon/{pokemon_id}/encounters", headers={"User-Agent": "RouteDex/0.1"})
    try:
        data = json.loads(urllib.request.urlopen(request, timeout=20).read().decode("utf-8"))
    except Exception:
        data = []
    cache.write_text(json.dumps(data), encoding="utf-8")
    return data


def api_location_name(value: str):
    value = re.sub(r"-area(-\d+)?$", "", value)
    value = re.sub(r"-\d+$", "", value)
    aliases = {
        "mt-coronet": "Mt. Coronet",
        "oreburgh-mine": "Oreburgh Mine",
        "ravaged-path": "Ravaged Path",
        "wayward-cave": "Wayward Cave",
        "lost-tower": "Lost Tower",
        "pokemon-mansion": "Pokemon Mansion",
        "great-marsh": "Great Marsh",
        "trophy-garden": "Trophy Garden",
        "old-chateau": "Old Chateau",
        "iron-island": "Iron Island",
        "ramanas-park": "Ramanas Park",
    }
    if value in aliases:
        return aliases[value]
    return re.sub(r"\b\w", lambda match: match.group(0).upper(), value.replace("-", " "))


def enrich_times(pokemon_ids):
    result = {}
    with ThreadPoolExecutor(max_workers=12) as executor:
        futures = {executor.submit(api_encounters, pokemon_id): pokemon_id for pokemon_id in pokemon_ids}
        for future in as_completed(futures):
            pokemon_id = futures[future]
            entries = []
            for area in future.result():
                area_name = api_location_name(area.get("location_area", {}).get("name", ""))
                for version in area.get("version_details", []):
                    if version.get("version", {}).get("name") not in {"diamond", "pearl"}:
                        continue
                    for detail in version.get("encounter_details", []):
                        times = []
                        for condition in detail.get("condition_values", []):
                            name = condition.get("name", "")
                            if name == "time-morning": times.append("morning")
                            if name == "time-day": times.append("day")
                            if name == "time-night": times.append("night")
                        entries.append({"location": area_name, "method": detail.get("method", {}).get("name"), "times": times or ["unknown"], "version": version.get("version", {}).get("name"), "source": "PokéAPI encounter conditions (D/P baseline)"})
            result[pokemon_id] = entries
    return result


def canonical_name(value: str) -> str:
    return clean(value).lower().replace("♀", "-f").replace("♂", "-m").replace("’", "'")


def pokemondb_method(title: str) -> str:
    normalized = title.lower().replace("é", "e")
    if normalized == "walking": return "Walking"
    if "surf" in normalized: return "Surf"
    if "old rod" in normalized: return "Old Rod"
    if "good rod" in normalized: return "Good Rod"
    if "super rod" in normalized: return "Super Rod"
    if "pokeradar" in normalized or "poke radar" in normalized: return "Poké Radar"
    if "swarm" in normalized: return "Swarm"
    return title.strip()


def pokemondb_slug(location_name: str) -> str | None:
    if location_name in {"Grand Underground", "Evolution", "Starter Pokémon", "Honey Trees", "Other routes"}:
        return None
    value = re.sub(r"[^a-zA-Z0-9]+", "-", location_name).strip("-").lower()
    aliases = {"mt-coronet": "mt-coronet", "pokemon-league": "pokemon-league", "pokemon-mansion": "pokemon-mansion", "old-chateau": "old-chateau"}
    return f"sinnoh-{aliases.get(value, value)}"


def parse_pokemondb_location(location):
    slug_value = pokemondb_slug(location["name"])
    if not slug_value:
        return {}
    POKEMONDB_CACHE.mkdir(parents=True, exist_ok=True)
    cache = POKEMONDB_CACHE / f"{location['id']}.html"
    try:
        if cache.exists():
            raw = cache.read_bytes()
        else:
            url = f"https://pokemondb.net/location/{slug_value}"
            request = urllib.request.Request(url, headers={"User-Agent": "RouteDex/0.1"})
            raw = urllib.request.urlopen(request, timeout=20).read()
            cache.write_bytes(raw)
        tree = lxml_html.fromstring(raw)
        heading = [node for node in tree.xpath("//h2") if "generation 8" in " ".join(node.text_content().lower().split())]
        if not heading:
            return {}
        details = {}
        method = "Walking"
        for node in heading[0].xpath("following::*"):
            if node.tag == "h2" and "generation 4" in " ".join(node.text_content().lower().split()):
                break
            if node.tag == "h3":
                method = pokemondb_method(node.text_content())
            if node.tag != "table":
                continue
            for row in node.xpath(".//tr"):
                cells = row.xpath("./td")
                if not cells:
                    continue
                name_nodes = cells[0].xpath('.//a[contains(@class,"ent-name")]')
                if not name_nodes:
                    continue
                name = " ".join(name_nodes[0].text_content().split())
                game_cells = row.xpath('.//td[contains(@class,"cell-loc-game")]')
                game_names = [" ".join(game.text_content().split()) for game in game_cells]
                versions = ["BD" if game_name == "BD" else "SP" for game_name in game_names if game_name in {"BD", "SP"}]
                if not versions:
                    continue
                time_cells = row.xpath('.//td[contains(@class,"cell-fixed") and not(contains(@class,"cell-name"))]')
                times = []
                if time_cells:
                    for image in time_cells[0].xpath(".//img"):
                        alt = image.get("alt", "").lower()
                        if alt == "morning": times.append("morning")
                        if alt == "day": times.append("day")
                        if alt == "night": times.append("night")
                if not times:
                    times = ["unknown"]
                numeric_cells = [" ".join(cell.text_content().split()) for cell in row.xpath('.//td[contains(@class,"cell-num")]')]
                chance = next((value for value in numeric_cells if "%" in value), None)
                levels = numeric_cells[-1] if numeric_cells and numeric_cells[-1] != chance else None
                rarity_nodes = row.xpath('.//span[contains(@class,"icon-rarity")]')
                rarity = " ".join(rarity_nodes[0].text_content().split()) if rarity_nodes else None
                detail_cells = row.xpath("./td")
                detail_text = " ".join(detail_cells[-1].text_content().split()) if detail_cells else ""
                entry = {"method": method, "times": times, "versions": versions, "chance": chance, "rarity": rarity, "levels": levels, "condition": detail_text or None, "source": f"https://pokemondb.net/location/{slug_value}"}
                details.setdefault(canonical_name(name), []).append(entry)
        return details
    except Exception:
        return {}


def enrich_from_pokemondb(locations):
    result = {}
    with ThreadPoolExecutor(max_workers=8) as executor:
        futures = {executor.submit(parse_pokemondb_location, location): location for location in locations.values()}
        for future in as_completed(futures):
            location = futures[future]
            result[location["id"]] = future.result()
    return result


def main():
    workbook = openpyxl.load_workbook(SOURCE, data_only=True, read_only=True)
    national = workbook["National"]
    pokemon = []
    location_map = {}
    acquisitions = []
    corrections = []
    raw_locations = {}
    for row in national.iter_rows(min_row=2, values_only=True):
        number, name, types, found = int(row[0]), clean(row[1]), clean(row[2]), clean(row[3])
        name = name.replace("Kriketune", "Kricketune")
        pokemon.append({"id": number, "name": name, "types": [item for item in types.split("\n") if item]})
        raw_locations[number] = list(location_parts(found))
        for raw in raw_locations[number]:
            if is_acquisition(raw):
                acquisitions.append({"id": f"acquisition-{number}-{len(acquisitions)}", "pokemonId": number, "group": acquisition_group(raw), "label": raw, "source": "Cópia de BDSP Pokedex Worklist Sharable.xlsx"})
                continue
            location_name, category, method, condition, times = normalize_location(raw)
            if location_name != raw:
                corrections.append({"from": raw, "to": location_name})
            location_id = slug(location_name)
            location_map.setdefault(location_id, {"id": location_id, "name": location_name, "category": category, "encounters": []})
            initial_times = times or ["unknown"]
            location_map[location_id]["encounters"].append({"id": f"{location_id}-{number}", "pokemonId": number, "locationId": location_id, "method": method, "condition": condition, "times": initial_times, "versions": ["BD", "SP"], "details": [{"method": method or "Walking", "times": initial_times, "versions": ["BD", "SP"], "condition": condition, "source": "Cópia de BDSP Pokedex Worklist Sharable.xlsx"}], "source": "Cópia de BDSP Pokedex Worklist Sharable.xlsx"})

    sinnoh = workbook["Sinnoh"]
    sinnoh_by_name = {}
    for row in sinnoh.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            sinnoh_by_name[clean(row[1])] = int(row[0])
    for item in pokemon:
        if item["name"] in sinnoh_by_name:
            item["sinnohNumber"] = sinnoh_by_name[item["name"]]

    enriched = enrich_times([item["id"] for item in pokemon])
    pokemondb = enrich_from_pokemondb(location_map)
    for location in location_map.values():
        for encounter in location["encounters"]:
            pokemon_name = next((item["name"] for item in pokemon if item["id"] == encounter["pokemonId"]), "")
            original_method = encounter.get("method")
            database_details = pokemondb.get(location["id"], {}).get(canonical_name(pokemon_name), [])
            if database_details:
                if original_method == "Swarm":
                    # Pokémon Database exposes the regular grass table, but a
                    # Swarm is a separate post-game encounter condition. Keep
                    # the workbook's method and avoid relabeling it as Walking.
                    encounter["details"] = [{"method": "Swarm", "times": encounter["times"], "versions": encounter["versions"], "condition": encounter["condition"], "source": encounter["source"]}]
                else:
                    encounter["details"] = database_details
                    encounter["times"] = sorted({time for detail in database_details for time in detail["times"]})
                    encounter["versions"] = sorted({version for detail in database_details for version in detail["versions"]})
                    encounter["method"] = " · ".join(sorted({detail["method"] for detail in database_details}))
                    encounter["source"] = database_details[0]["source"]
            matches = [entry for entry in enriched.get(encounter["pokemonId"], []) if entry["location"].lower() == location["name"].lower()]
            times = sorted({time for match in matches for time in match["times"]})
            if times and encounter["times"] == ["unknown"]:
                encounter["times"] = times
            elif times:
                encounter["times"] = sorted(set(encounter["times"]) | set(times))
            matched_versions = sorted({"BD" if match.get("version") == "diamond" else "SP" for match in matches if match.get("version") in {"diamond", "pearl"}})
            if matched_versions:
                encounter["versions"] = matched_versions
            encounter["source"] = next((match["source"] for match in matches if match["source"]), encounter["source"])

    postgame = []
    for row in workbook["Postgame Checklist"].iter_rows(min_row=2, values_only=True):
        if row[1] is not None:
            postgame.append({"id": int(row[1]), "title": clean(row[2])})
    tms = []
    for row in workbook["TM Locations"].iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            tms.append({"id": int(row[0]), "name": clean(row[1]), "location": clean(row[2])})

    unique_corrections = []
    seen_corrections = set()
    for correction in corrections:
        key = (correction["from"], correction["to"])
        if key not in seen_corrections:
            seen_corrections.add(key)
            unique_corrections.append(correction)
    unknown_time_records = [{"location": location["name"], "pokemonId": encounter["pokemonId"]} for location in location_map.values() for encounter in location["encounters"] if "unknown" in encounter["times"]]
    data = {"pokemon": pokemon, "locations": sorted(location_map.values(), key=lambda item: (item["category"], item["name"])), "acquisitions": acquisitions, "postgame": postgame, "tms": tms, "quality": {"corrections": unique_corrections, "unknownTimeRecords": unknown_time_records}}
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text("// Generated by scripts/import_workbook.py\nexport const gameData = " + json.dumps(data, ensure_ascii=False, indent=2) + " as const;\n", encoding="utf-8")
    print(f"Generated {OUTPUT} with {len(pokemon)} Pokémon, {len(location_map)} locations, {len(acquisitions)} acquisitions, {len(postgame)} postgame items and {len(tms)} TMs")


if __name__ == "__main__":
    main()
